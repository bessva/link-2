# -*- coding: utf-8 -*-
"""
app.py — ЛИНК v5: Энергетический ассистент
Запуск: streamlit run app.py

Три режима:
  1. Анализ станций из базы (Excel)   — режим «Аналитик»
  2. Расчёт новой станции (диалог)    — режим «Новая станция»
  3. Ручной расчёт формулы            — режим «Калькулятор»

Зависимости модулей (кладите в ту же папку):
  calculations.py
  station_profile.py
  report_generator.py
"""

import os
import re
import uuid
import time
import glob
import requests
import streamlit as st
from datetime import datetime

# Наши модули
from calculations    import run_full_calculation
from station_profile import (
    new_profile, get_next_question, parse_user_answer,
    format_question, is_profile_complete, apply_auto_fields,
    apply_defaults, profile_summary_text
)
from report_generator import (
    build_report_markdown, build_gigachat_context,
    results_to_dataframe, results_to_excel_bytes
)

# ============================================================
# ⚙️ НАСТРОЙКИ
# ============================================================

GIGACHAT_AUTH_KEY = "MDE5Y2E1MTQtMzEzYS03NzRkLTk0NWEtYTk0YjU0ZGFlMmM3OmE2ODk0ZjViLTMyMmMtNDUwYy1hZTdiLTRlY2ZiOGMxMTdlNw=="

BASE_DIR      = os.path.dirname(os.path.abspath(__file__))
DATA_DIR      = os.path.join(BASE_DIR, "data")
KNOWLEDGE_DIR = os.path.join(BASE_DIR, "knowledge")
VECTOR_DB_DIR = os.path.join(BASE_DIR, "vector_db")

for d in [DATA_DIR, KNOWLEDGE_DIR, VECTOR_DB_DIR]:
    os.makedirs(d, exist_ok=True)


# ============================================================
# 🎨 СТИЛИ
# ============================================================

st.set_page_config(
    page_title="ЛИНК — Энергетический ассистент",
    page_icon="⚡",
    layout="wide",
    initial_sidebar_state="expanded"
)

st.markdown("""
<link href="https://fonts.googleapis.com/css2?family=Montserrat:wght@300;400;500;600&display=swap" rel="stylesheet">
<style>
.stApp { background-color: #0d1117; font-family: 'Montserrat', sans-serif; }
[data-testid="stSidebar"] { background-color: #161b22; border-right: 1px solid #21262d; }
[data-testid="stSidebar"] * { color: #c9d1d9 !important; }
.sidebar-title { font-family: 'Montserrat', monospace; font-size: 18px; font-weight: 600;
    color: #58a6ff !important; letter-spacing: 2px; text-transform: uppercase; margin-bottom: 4px; }
.sidebar-subtitle { font-size: 11px; color: #8b949e !important; letter-spacing: 1px;
    text-transform: uppercase; margin-bottom: 24px; }
.sidebar-section { font-size: 10px; font-weight: 600; letter-spacing: 2px;
    text-transform: uppercase; color: #58a6ff !important;
    margin: 20px 0 8px 0; padding-bottom: 4px; border-bottom: 1px solid #21262d; }
.sidebar-item { font-size: 12px; color: #8b949e !important; padding: 3px 0; }
.sidebar-item-active { color: #c9d1d9 !important; }
.main-header { font-family: 'Montserrat', monospace; font-size: 13px; color: #58a6ff;
    letter-spacing: 3px; text-transform: uppercase;
    padding: 16px 0 4px 0; border-bottom: 1px solid #21262d; margin-bottom: 24px; }
.msg-user { background: #1c2128; border: 1px solid #30363d;
    border-radius: 8px 8px 2px 8px; padding: 12px 16px;
    margin-left: 60px; color: #c9d1d9; font-size: 14px; line-height: 1.6; }
.msg-assistant { background: #161b22; border: 1px solid #21262d;
    border-left: 3px solid #58a6ff; border-radius: 2px 8px 8px 8px;
    padding: 12px 16px; margin-right: 60px; color: #c9d1d9; font-size: 14px; line-height: 1.6; }
.msg-label-user { font-size: 10px; color: #8b949e; letter-spacing: 1px;
    text-transform: uppercase; margin-bottom: 6px; text-align: right; }
.msg-label-assistant { font-size: 10px; color: #58a6ff; letter-spacing: 1px;
    text-transform: uppercase; margin-bottom: 6px; font-family: 'Montserrat', monospace; }
.msg-system { background: #1a2332; border: 1px dashed #30363d; border-radius: 6px;
    padding: 10px 14px; color: #8b949e; font-size: 13px; line-height: 1.5;
    margin: 4px 0; font-style: italic; }
.source-badge { display: inline-block; font-size: 9px; letter-spacing: 1px;
    text-transform: uppercase; padding: 2px 6px; border-radius: 3px;
    margin-right: 4px; font-family: 'Montserrat', monospace; }
.source-excel { background: #1a3a1a; color: #3fb950; border: 1px solid #238636; }
.source-kb    { background: #1a1f3a; color: #79c0ff; border: 1px solid #1f6feb; }
.source-calc  { background: #2d2a1a; color: #f0c040; border: 1px solid #9e7c00; }
.source-new   { background: #2a1a3a; color: #d2a8ff; border: 1px solid #6e40c9; }
.stTextInput input { background-color: #161b22 !important; border: 1px solid #30363d !important;
    border-radius: 6px !important; color: #c9d1d9 !important;
    font-family: 'Montserrat', sans-serif !important; font-size: 14px !important;
    padding: 10px 14px !important; }
.stTextInput input:focus { border-color: #58a6ff !important;
    box-shadow: 0 0 0 3px rgba(88,166,255,0.1) !important; }
.stButton button { background-color: #21262d !important; color: #c9d1d9 !important;
    border: 1px solid #30363d !important; border-radius: 6px !important;
    font-family: 'Montserrat', sans-serif !important; font-size: 11px !important;
    letter-spacing: 0.5px !important; transition: all 0.2s !important;
    height: 36px !important; padding: 0 8px !important; }
.stButton button:hover { background-color: #30363d !important;
    border-color: #58a6ff !important; color: #58a6ff !important; }
[data-testid="stSidebar"] .stButton button,
[data-testid="stSidebar"] .stButton button p {
    font-size: 12px !important;
    height: auto !important;
    padding: 5px 4px !important;
    white-space: normal !important;
    line-height: 1.5 !important; }
[data-testid="stForm"] [data-testid="stFormSubmitButton"]:first-child button {
    background-color: #1f6feb !important; color: #ffffff !important;
    border-color: #58a6ff !important; }
[data-testid="stForm"] [data-testid="stFormSubmitButton"]:first-child button:hover {
    background-color: #388bfd !important; border-color: #79c0ff !important; }
[data-testid="stForm"] [data-testid="stFormSubmitButton"]:last-child button {
    background-color: rgba(33, 38, 45, 0.4) !important; color: #8b949e !important;
    border-color: #30363d !important; }
[data-testid="stForm"] [data-testid="stFormSubmitButton"]:last-child button:hover {
    background-color: rgba(48, 54, 61, 0.6) !important; color: #c9d1d9 !important; }
.mode-card { background: #161b22; border: 1px solid #30363d; border-radius: 8px;
    padding: 16px; margin: 8px 0; cursor: pointer; transition: all 0.2s; }
.mode-card:hover { border-color: #58a6ff; }
.status-ok  { color: #3fb950; font-size: 11px; }
.status-err { color: #f85149; font-size: 11px; }
#MainMenu, footer, header { visibility: hidden; }
.block-container { padding-top: 20px !important; }
</style>
""", unsafe_allow_html=True)


# ============================================================
# 🔐 GIGACHAT
# ============================================================

def get_access_token():
    if ("_token" in st.session_state and
            time.time() < st.session_state.get("_token_exp", 0) - 60):
        return st.session_state["_token"]
    url = "https://ngw.devices.sberbank.ru:9443/api/v2/oauth"
    headers = {
        "Content-Type":  "application/x-www-form-urlencoded",
        "Accept":        "application/json",
        "RqUID":         str(uuid.uuid4()),
        "Authorization": f"Basic {GIGACHAT_AUTH_KEY}",
    }
    r = requests.post(url, headers=headers, data={"scope": "GIGACHAT_API_PERS"},
                      verify=False, timeout=15)
    r.raise_for_status()
    d = r.json()
    st.session_state["_token"]     = d["access_token"]
    st.session_state["_token_exp"] = d.get("expires_at", 0) / 1000
    return st.session_state["_token"]


def call_gigachat(system_prompt: str, user_message: str, history: list) -> str:
    token    = get_access_token()
    messages = [{"role": "system", "content": system_prompt}]
    for q, a, _ in history[-6:]:
        messages.append({"role": "user",      "content": q})
        messages.append({"role": "assistant", "content": a})
    messages.append({"role": "user", "content": user_message})
    r = requests.post(
        "https://gigachat.devices.sberbank.ru/api/v1/chat/completions",
        headers={"Accept": "application/json", "Content-Type": "application/json",
                 "Authorization": f"Bearer {token}"},
        json={"model": "GigaChat-Pro", "messages": messages, "temperature": 0.3, "max_tokens": 1500},
        verify=False, timeout=30
    )
    r.raise_for_status()
    return r.json()["choices"][0]["message"]["content"].strip()


# ============================================================
# 📊 EXCEL — ЗАГРУЗКА ДАННЫХ ПО СТАНЦИЯМ
# ============================================================

def _find_col(headers, keywords):
    for i, h in enumerate(headers):
        if h and any(kw in str(h).lower() for kw in keywords):
            return i
    return None


@st.cache_data(show_spinner=False)
def load_excel_data(folder: str) -> dict:
    try:
        import openpyxl
    except ImportError:
        return {}
    all_data = {}
    for filepath in glob.glob(os.path.join(folder, "*.xlsx")):
        try:
            wb = openpyxl.load_workbook(filepath, data_only=True)
            for sheet_name in wb.sheetnames:
                ws   = wb[sheet_name]
                rows = list(ws.iter_rows(values_only=True))
                if len(rows) < 3:
                    continue
                headers    = rows[0]
                col_tes    = _find_col(headers, ["наименование тэс", "наименование"])
                col_month  = _find_col(headers, ["месяц"])
                col_fuel   = _find_col(headers, ["вид топлива"])
                sheet_up   = sheet_name.upper()
                result_col = None
                if   "ННЗТ" in sheet_up: result_col = _find_col(headers, ["ннзт"])
                elif "НЭЗТ" in sheet_up: result_col = _find_col(headers, ["нэзт, т.н.т", "нэзт"])
                elif "НАЗТ" in sheet_up: result_col = _find_col(headers, ["назт, т.н.т", "назт"])
                elif "ОНЗТ" in sheet_up: result_col = _find_col(headers, ["онзт, т.н.т", "онзт"])
                if result_col is None or col_month is None:
                    continue
                current_tes = current_fuel = None
                for row in rows[2:]:
                    if col_tes  is not None and row[col_tes]:  current_tes  = str(row[col_tes]).strip()
                    if col_fuel is not None and row[col_fuel]: current_fuel = str(row[col_fuel]).strip()
                    month_val = row[col_month] if col_month is not None else None
                    result    = row[result_col]
                    if not current_tes or not month_val or result is None:
                        continue
                    (all_data
                     .setdefault(current_tes, {})
                     .setdefault(sheet_up, {})
                     [str(month_val).strip()]) = {
                        "value": round(float(result), 3),
                        "fuel":  current_fuel or "—"
                    }
        except Exception:
            pass
    return all_data


def find_relevant_excel(excel_data: dict, question: str) -> str:
    q = question.lower()
    norm_types = [n for n in ["ННЗТ","НЭЗТ","НАЗТ","ОНЗТ"] if n.lower() in q] \
                 or ["ННЗТ","НЭЗТ","НАЗТ","ОНЗТ"]
    months_ru  = ["январь","февраль","март","апрель","май","июнь",
                  "июль","август","сентябрь","октябрь","ноябрь","декабрь"]
    target_months   = [m.capitalize() for m in months_ru if m in q]
    target_stations = [t for t in excel_data
                       if any(w[:5] in q for w in t.lower().split() if len(w) > 4)] \
                      or list(excel_data.keys())
    lines = []
    for tes in target_stations:
        block = [f"📍 {tes}:"]
        for norm in norm_types:
            if norm in excel_data.get(tes, {}):
                block.append(f"  {norm}:")
                for month, info in excel_data[tes][norm].items():
                    if not target_months or month in target_months:
                        block.append(f"    {month}: {info['value']} т.н.т.")
        if len(block) > 1:
            lines.extend(block)
    return "\n".join(lines)


# ============================================================
# 📚 БАЗА ЗНАНИЙ (ChromaDB + SentenceTransformers)
# ============================================================

def extract_pdf(path: str) -> str:
    try:
        import pdfplumber
        text = ""
        with pdfplumber.open(path) as pdf:
            for page in pdf.pages:
                t = page.extract_text()
                if t: text += t + "\n"
        return text
    except Exception:
        return ""


def extract_docx(path: str) -> str:
    try:
        from docx import Document
        doc = Document(path)
        return "\n".join(p.text for p in doc.paragraphs if p.text.strip())
    except Exception:
        return ""


def chunk_text(text: str, size: int = 800, overlap: int = 100) -> list:
    words  = text.split()
    chunks, i = [], 0
    while i < len(words):
        c = " ".join(words[i:i+size])
        if c.strip(): chunks.append(c)
        i += size - overlap
    return chunks


@st.cache_resource(show_spinner=False)
def build_knowledge_base(knowledge_dir: str, vector_db_dir: str):
    try:
        import chromadb
        from sentence_transformers import SentenceTransformer
    except ImportError:
        return None, None
    files = (glob.glob(os.path.join(knowledge_dir, "*.pdf")) +
             glob.glob(os.path.join(knowledge_dir, "*.docx")))
    if not files:
        return None, None
    client   = chromadb.PersistentClient(path=vector_db_dir)
    existing = [c.name for c in client.list_collections()]
    model    = SentenceTransformer("paraphrase-multilingual-mpnet-base-v2")
    if "knowledge" in existing:
        col = client.get_collection("knowledge")
        if col.count() > 0:
            return col, model
    if "knowledge" in existing:
        client.delete_collection("knowledge")
    col = client.create_collection("knowledge")
    all_chunks, all_ids, all_metas = [], [], []
    for filepath in files:
        fname = os.path.basename(filepath)
        ext   = os.path.splitext(filepath)[1].lower()
        text  = extract_pdf(filepath) if ext == ".pdf" else extract_docx(filepath)
        if not text.strip(): continue
        for i, chunk in enumerate(chunk_text(text)):
            all_chunks.append(chunk)
            all_ids.append(f"{fname}_{i}")
            all_metas.append({"source": fname, "chunk": i})
    if all_chunks:
        embeddings = model.encode(all_chunks, show_progress_bar=False).tolist()
        col.add(documents=all_chunks, embeddings=embeddings,
                ids=all_ids, metadatas=all_metas)
    return col, model


def search_kb(question: str, collection, model, top_k: int = 3) -> str:
    if collection is None or model is None: return ""
    try:
        q_emb   = model.encode([question]).tolist()
        results = collection.query(query_embeddings=q_emb, n_results=top_k)
        docs    = results.get("documents", [[]])[0]
        metas   = results.get("metadatas", [[]])[0]
        return "\n\n---\n\n".join(
            f"[{m.get('source','?')}]\n{d}" for d, m in zip(docs, metas)
        )
    except Exception:
        return ""


# ============================================================
# 🧮 КАЛЬКУЛЯТОР — РУЧНОЙ РАСЧЁТ ОДНОЙ ФОРМУЛЫ
# ============================================================

CALC_FORMULAS = {
    "nnzt_pipeline": {
        "label":   "ННЗТ — трубопровод/газ (3 сут)",
        "keywords": ["ннзт", "неснижаем", "труба", "газопровод"],
        "formula": "ННЗТ = 3 × Bусл × 7000 / Qнр",
        "params":  ["B_usl", "Q_nr"],
        "prompts": {
            "B_usl": "Bусл — суточный расход условного топлива в режиме выживания (т.у.т./сут)",
            "Q_nr":  "Qнр — теплота сгорания натурального топлива (ккал/кг)",
        },
        "calc": lambda B_usl, Q_nr: 3 * B_usl * 7000 / Q_nr,
        "units": "т.н.т.",
    },
    "nnzt_solid": {
        "label":   "ННЗТ — уголь/мазут/дизель (7 сут)",
        "keywords": ["ннзт", "неснижаем", "уголь", "мазут", "дизель"],
        "formula": "ННЗТ = 7 × Bусл × 7000 / Qнр",
        "params":  ["B_usl", "Q_nr"],
        "prompts": {
            "B_usl": "Bусл — суточный расход условного топлива в режиме выживания (т.у.т./сут)",
            "Q_nr":  "Qнр — теплота сгорания натурального топлива (ккал/кг)",
        },
        "calc": lambda B_usl, Q_nr: 7 * B_usl * 7000 / Q_nr,
        "units": "т.н.т.",
    },
    "nezt_small": {
        "label":   "НЭЗТ — малые ТЭС (<25 МВт)",
        "keywords": ["нэзт", "менее 25", "<25", "малая"],
        "formula": "НЭЗТ = Bрср × Тпср × Кср",
        "params":  ["B_rsr", "T_psr", "K_sr"],
        "prompts": {
            "B_rsr": "Bрср — среднерасчётный суточный расход топлива (т/сут)",
            "T_psr": "Тпср — среднее время поставки (сут)",
            "K_sr":  "Кср — коэффициент срыва поставок (1.2–3.5)",
        },
        "calc": lambda B_rsr, T_psr, K_sr: B_rsr * T_psr * K_sr / 1000,
        "units": "тыс.т.н.т.",
    },
    "nezt_large_pipeline": {
        "label":   "НЭЗТ — газ/трубопровод (≥25 МВт, Ф.29)",
        "keywords": ["нэзт", "газ", "более 25", "≥25", "трубопровод"],
        "formula": "НЭЗТ = НЭЗТб.в. × RТЭС",
        "params":  ["NEZT_bv", "R_tes"],
        "prompts": {
            "NEZT_bv": "НЭЗТб.в. — базовая величина НЭЗТ (т.н.т.)",
            "R_tes":   "RТЭС — коэффициент риска (0–1; очень высокий=1, средний=0.5, очень низкий=0)",
        },
        "calc": lambda NEZT_bv, R_tes: NEZT_bv * R_tes,
        "units": "т.н.т.",
    },
    "nezt_large_solid": {
        "label":   "НЭЗТ — уголь/торф/дизель (≥25 МВт, Ф.29(3))",
        "keywords": ["нэзт", "уголь", "торф", "твёрдое", "более 25"],
        "formula": "НЭЗТ = НЭЗТб.в. × Кпост × Кср",
        "params":  ["NEZT_bv", "K_post", "K_sr"],
        "prompts": {
            "NEZT_bv": "НЭЗТб.в. — базовая величина НЭЗТ (т.н.т.)",
            "K_post":  "Кпост — коэффициент времени поставки (0.3–1.8, по таблице Т)",
            "K_sr":    "Кср — коэффициент срыва поставки (1.0–2.5)",
        },
        "calc": lambda NEZT_bv, K_post, K_sr: NEZT_bv * K_post * K_sr,
        "units": "т.н.т.",
    },
    "nazt_small": {
        "label":   "НАЗТ — малые ТЭС с ПГУ/ГТУ (Ф.18)",
        "keywords": ["назт", "аварийный", "пгу", "гту"],
        "formula": "НАЗТ = Bсут × N × k / 24",
        "params":  ["B_sut", "N", "k"],
        "prompts": {
            "B_sut": "Bсут — суточный расход аварийного топлива (т/сут)",
            "N":     "N — суток работы на аварийном топливе (3–5, по проектной документации)",
            "k":     "k — часов работы (обычно 24)",
        },
        "calc": lambda B_sut, N, k: B_sut * N * k / 24 / 1000,
        "units": "тыс.т.н.т.",
    },
    "nazt_large": {
        "label":   "НАЗТ — крупные ТЭС с ПГУ/ГТУ (Ф.36–37)",
        "keywords": ["назт", "аварийный", "пгу", "гту", "более 25"],
        "formula": "НАЗТ = 3 × Bсут × RТЭС (не менее 2/3 × НАЗТб.в.)",
        "params":  ["B_sut", "R_tes"],
        "prompts": {
            "B_sut": "Bсут — суточный расход аварийного топлива (т/сут)",
            "R_tes": "RТЭС — коэффициент риска (0–1)",
        },
        "calc": lambda B_sut, R_tes: max(3 * B_sut * R_tes, (2/3) * 3 * B_sut),
        "units": "т.н.т.",
    },
    "nezt_bv": {
        "label":   "НЭЗТб.в. — базовая величина (Ф.25)",
        "keywords": ["нэзт б.в.", "базовая величина", "нэзт бв"],
        "formula": "НЭЗТб.в. = Bмакс × nсут × 7000 / Qн",
        "params":  ["B_maks", "n_sut", "Q_n"],
        "prompts": {
            "B_maks": "Bмакс — максимальный суточный расход условного топлива (т.у.т./сут)",
            "n_sut":  "nсут — 3 (газ/трубопровод) или 7 (уголь и т.д.)",
            "Q_n":    "Qн — теплота сгорания натурального топлива (ккал/кг)",
        },
        "calc": lambda B_maks, n_sut, Q_n: B_maks * n_sut * 7000 / Q_n,
        "units": "т.н.т.",
    },
    "kium_t": {
        "label":   "КИУМт — коэффициент использования тепловой мощности (Ф.40)",
        "keywords": ["киум", "коэффициент использования", "тепловая мощность"],
        "formula": "КИУМт = Qтс_план / (Qтс_уст × nмес) × 100%",
        "params":  ["Q_plan", "Q_ust", "n_mes"],
        "prompts": {
            "Q_plan": "Qтс_план — плановый отпуск тепла (Гкал)",
            "Q_ust":  "Qтс_уст — установленная тепловая мощность (Гкал/ч) × 720 часов в месяц",
            "n_mes":  "nмес — число месяцев (обычно 1)",
        },
        "calc": lambda Q_plan, Q_ust, n_mes: (Q_plan / (Q_ust * n_mes)) * 100 if Q_ust * n_mes > 0 else 0,
        "units": "%",
    },
}

CALC_TRIGGER_WORDS = [
    "рассчитай", "посчитай", "вычисли", "расчёт", "расчет",
    "калькулятор", "формула", "по формуле", "сколько будет"
]


def detect_calc_formula(user_input: str) -> tuple:
    """
    Возвращает (formula_key, formula_data) или (None, None).
    Срабатывает только если есть слово-триггер расчёта.
    """
    q = user_input.lower()
    if not any(kw in q for kw in CALC_TRIGGER_WORDS):
        return None, None
    priority = [
        "kium_t", "nazt_large", "nazt_small", "nezt_bv",
        "nnzt_pipeline", "nnzt_solid",
        "nezt_large_pipeline", "nezt_large_solid", "nezt_small",
    ]
    for key in priority:
        data = CALC_FORMULAS[key]
        if any(kw in q for kw in data["keywords"]):
            return key, data
    return None, None
 
MODE_SWITCH_COMMANDS = {
    "analyst":     ["база", "анализ", "аналитик", "данные станций", "из базы"],
    "new_station": ["новая станция", "новый расчёт", "рассчитать станцию", "начать расчёт"],
    "calc":        ["калькулятор", "ручной расчёт", "одна формула"],
}

def detect_mode_switch(user_input: str) -> str | None:
    q = user_input.lower().strip()
    # Универсальные команды выхода
    if q in ("меню", "старт", "назад", "смена режима", "сменить режим", "выход"):
        return "menu"
    for mode, keywords in MODE_SWITCH_COMMANDS.items():
        if any(kw in q for kw in keywords):
            return mode
    return None


# ============================================================
# 🧠 СИСТЕМНЫЕ ПРОМПТЫ GIGACHAT
# ============================================================

SYSTEM_PROMPT_ANALYST = (
    "Ты — эксперт-аналитик по управлению энергетическими объектами и нормативам "
    "запасов топлива (ННЗТ, НЭЗТ, НАЗТ, ОНЗТ, НВЗТ) согласно Приказу Минэнерго №1062. "
    "Анализируй числовые данные по станциям, давай управленческие рекомендации. "
    "Отвечай кратко: 3–5 предложений или короткий список. Только факты и выводы. "
    "Отвечай ТОЛЬКО на вопросы про энергетику, ТЭС, топливные запасы. "
    "На остальные темы: 'Я специализируюсь на энергетике.'"
)

SYSTEM_PROMPT_STATION = (
    "Ты — эксперт по нормативам топливных запасов ТЭС (Приказ Минэнерго №1062). "
    "У тебя есть полные рассчитанные данные по конкретной станции. "
    "Отвечай на вопросы ТОЛЬКО на основе этих данных. "
    "Если данных нет — скажи об этом прямо. "
    "Отвечай кратко и по делу."
)


# ============================================================
# 🖥️ SESSION STATE
# ============================================================

def init_session():
    defaults = {
        "history":         [],       # [(user_msg, bot_msg, source), ...]
        "mode":            None,     # None | "analyst" | "new_station" | "calc"
        "calc_state":      None,     # {key, data, collected_params}
        "station_profile": None,     # dict (текущий профиль в сборке)
        "station_results": None,     # dict (результаты расчёта)
        "station_context": None,     # str (контекст для GigaChat)
    }
    for k, v in defaults.items():
        if k not in st.session_state:
            st.session_state[k] = v


init_session()

excel_data = load_excel_data(DATA_DIR)

with st.spinner("Загружаю базу знаний..."):
    collection, st_model = build_knowledge_base(KNOWLEDGE_DIR, VECTOR_DB_DIR)


# ============================================================
# 🗂️ САЙДБАР
# ============================================================

with st.sidebar:
    st.markdown('<div class="sidebar-title">⚡ ЛИНК</div>', unsafe_allow_html=True)
    st.markdown('<div class="sidebar-subtitle">Энергетический ассистент</div>', unsafe_allow_html=True)

    # Текущий режим
    current_mode = st.session_state.mode
    mode_labels  = {
        None:          "— не выбран —",
        "analyst":     "📊 Анализ базы",
        "new_station": "🏭 Новая станция",
        "calc":        "📐 Калькулятор",
    }
    st.markdown('<div class="sidebar-section">Режим</div>', unsafe_allow_html=True)
    st.markdown(f'<div class="sidebar-item sidebar-item-active">{mode_labels.get(current_mode, "—")}</div>',
                unsafe_allow_html=True)
    st.markdown('<div class="sidebar-section">Сменить режим</div>', unsafe_allow_html=True)
    col_s1, col_s2 = st.sidebar.columns(2)
    with col_s1:
        if st.button("📊 База", use_container_width=True):
            st.session_state.mode = "analyst"
            st.session_state.calc_state = None
            st.session_state.history.append(("📊 Анализ станций из базы",
                f"Переключился в режим анализа. Загружено {len(excel_data)} станций.",
                "system"))
            st.rerun()
        if st.button("📐 Калькулятор", use_container_width=True):
            st.session_state.mode = "calc"
            st.session_state.calc_state = None
            st.session_state.history.append(("📐 Ручной расчёт",
                "Переключился в режим калькулятора.\n"
                "Напишите что хотите рассчитать, например: «рассчитай ННЗТ для газовой станции»",
                "system"))
            st.rerun()
    with col_s2:
        if st.button("🏭 Новая станция", use_container_width=True):
            st.session_state.mode = "new_station"
            st.session_state.station_profile = new_profile()
            st.session_state.calc_state = None
            first_step = get_next_question(st.session_state.station_profile)
            answer = ("Переключился в режим расчёта новой станции.\n\n" +
                      format_question(first_step)) if first_step else "—"
            st.session_state.history.append(("🏭 Новая станция", answer, "new_station"))
            st.rerun()

    # Станции из базы
    st.markdown('<div class="sidebar-section">Станции в базе</div>', unsafe_allow_html=True)
    if excel_data:
        for tes in list(excel_data.keys())[:10]:
            st.markdown(f'<div class="sidebar-item sidebar-item-active">🏭 {tes}</div>',
                        unsafe_allow_html=True)
    else:
        st.markdown('<div class="sidebar-item">Нет данных в папке data/</div>', unsafe_allow_html=True)

    # Текущая рассчитанная станция
    if st.session_state.station_results:
        st.markdown('<div class="sidebar-section">Рассчитанная станция</div>', unsafe_allow_html=True)
        name = st.session_state.station_results.get("station_name", "—")
        st.markdown(f'<div class="sidebar-item sidebar-item-active">✅ {name}</div>',
                    unsafe_allow_html=True)
        # Кнопка скачать Excel
        xls_bytes = results_to_excel_bytes(
            st.session_state.station_profile or {},
            st.session_state.station_results
        )
        if xls_bytes:
            st.download_button(
                label="📥 Скачать отчёт .xlsx",
                data=xls_bytes,
                file_name=f"normativy_{name.replace(' ','_')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

    # База знаний
    st.markdown('<div class="sidebar-section">База знаний</div>', unsafe_allow_html=True)
    kb_files = (glob.glob(os.path.join(KNOWLEDGE_DIR, "*.pdf")) +
                glob.glob(os.path.join(KNOWLEDGE_DIR, "*.docx")))
    if kb_files:
        for f in kb_files[:5]:
            st.markdown(f'<div class="sidebar-item sidebar-item-active">📄 {os.path.basename(f)}</div>',
                        unsafe_allow_html=True)
    else:
        st.markdown('<div class="sidebar-item">Нет документов в knowledge/</div>', unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)
    if st.button("🗑  Очистить чат"):
        for k in ["history", "mode", "calc_state", "station_profile",
                  "station_results", "station_context"]:
            st.session_state[k] = None if k != "history" else []
        st.rerun()

    st.markdown('<div class="sidebar-section">Статус</div>', unsafe_allow_html=True)
    try:
        get_access_token()
        st.markdown('<div class="status-ok">● GigaChat подключён</div>', unsafe_allow_html=True)
    except Exception:
        st.markdown('<div class="status-err">● Ошибка GigaChat</div>', unsafe_allow_html=True)

    if collection is not None:
        st.markdown(f'<div class="status-ok">● База знаний ({collection.count()} фрагм.)</div>',
                    unsafe_allow_html=True)
    else:
        st.markdown('<div class="status-err">● База знаний не загружена</div>', unsafe_allow_html=True)


# ============================================================
# 🖥️ ОСНОВНАЯ ОБЛАСТЬ
# ============================================================

st.markdown('<div class="main-header">// Диалог с ассистентом</div>', unsafe_allow_html=True)


def render_chat():
    """Отображает историю диалога."""
    source_badges = {
        "excel":       '<span class="source-badge source-excel">📊 Данные</span>',
        "knowledge":   '<span class="source-badge source-kb">📚 База знаний</span>',
        "both":        '<span class="source-badge source-excel">📊</span><span class="source-badge source-kb">📚</span>',
        "calc":        '<span class="source-badge source-calc">📐 Калькулятор</span>',
        "new_station": '<span class="source-badge source-new">🏭 Новая станция</span>',
        "system":      "",
    }
    for user_msg, bot_msg, source in st.session_state.history:
        badge = source_badges.get(source, "")
        st.markdown(f'<div class="msg-label-user">Вы</div>'
                    f'<div class="msg-user">{user_msg}</div>', unsafe_allow_html=True)
        st.markdown(f'<div class="msg-label-assistant">⚡ ЛИНК {badge}</div>'
                    f'<div class="msg-assistant">{bot_msg}</div>', unsafe_allow_html=True)
        st.markdown("<br>", unsafe_allow_html=True)


render_chat()

# ── Выбор режима (если не выбран) ────────────────────────

if st.session_state.mode is None:
    st.markdown('<div class="msg-system">'
                'Выберите режим работы или просто напишите вопрос:<br>'
                '<b>📊 Анализ станций из базы</b> — поиск и анализ данных из Excel<br>'
                '<b>🏭 Новая станция</b> — диалоговый сбор данных и расчёт нормативов<br>'
                '<b>📐 Калькулятор</b> — ручной расчёт одной формулы<br>'
                '</div>', unsafe_allow_html=True)

    col1, col2, col3 = st.columns(3)
    with col1:
        if st.button("📊 Анализ станций из базы", use_container_width=True):
            st.session_state.mode = "analyst"
            st.session_state.history.append((
                "📊 Анализ станций из базы",
                f"Загружено {len(excel_data)} станций. Задавайте вопросы по данным из базы.",
                "system"
            ))
            st.rerun()
    with col2:
        if st.button("🏭 Рассчитать новую станцию", use_container_width=True):
            st.session_state.mode = "new_station"
            st.session_state.station_profile = new_profile()
            first_step = get_next_question(st.session_state.station_profile)
            answer = ("Запускаю сбор данных по станции.\n\n" +
                      format_question(first_step)) if first_step else "Профиль уже заполнен."
            st.session_state.history.append(("🏭 Рассчитать новую станцию", answer, "new_station"))
            st.rerun()
    with col3:
        if st.button("📐 Ручной расчёт формулы", use_container_width=True):
            st.session_state.mode = "calc"
            st.session_state.history.append((
                "📐 Ручной расчёт формулы",
                "Напишите что хотите рассчитать, например:\n"
                "«рассчитай ННЗТ для угольной станции» или «посчитай НАЗТ»",
                "system"
            ))
            st.rerun()

st.markdown("<br>", unsafe_allow_html=True)

# ── Поле ввода ────────────────────────────────────────────

with st.form("chat_form", clear_on_submit=True):
    col_in, col_btn, col_skip = st.columns([5, 1, 1])
    with col_in:
        user_input = st.text_input(
            "",
            placeholder="Введите ответ или вопрос...",
            label_visibility="collapsed"
        )
    with col_btn:
        submitted = st.form_submit_button("Отправить", use_container_width=True)
    with col_skip:
        skipped = st.form_submit_button("Пропустить", use_container_width=True)


# ============================================================
# 🔄 ОБРАБОТКА ВВОДА
# ============================================================

if (submitted and user_input.strip()) or (skipped and st.session_state.mode == "new_station"):
    mode = st.session_state.mode

    # Проверяем команду смены режима
    switch_to = detect_mode_switch(user_input)
    if switch_to == "menu":
        st.session_state.mode = None
        st.session_state.calc_state = None
        st.session_state.history.append((user_input,
            "Возвращаю в главное меню. Выберите режим.",
            "system"))
        st.rerun()
    elif switch_to is not None:
        st.session_state.mode = switch_to
        st.session_state.calc_state = None
        if switch_to == "new_station":
            st.session_state.station_profile = new_profile()
            first_step = get_next_question(st.session_state.station_profile)
            answer = "Переключился в режим расчёта новой станции.\n\n" + format_question(first_step)
        elif switch_to == "analyst":
            answer = f"Переключился в режим анализа. Загружено {len(excel_data)} станций."
        elif switch_to == "calc":
            answer = "Переключился в режим калькулятора. Напишите что рассчитать."
        else:
            answer = "Режим переключён."
        st.session_state.history.append((user_input, answer, "system"))
        st.rerun()

    # Автоопределение режима если не выбран
    if mode is None:
        q = user_input.lower()
        if any(w in q for w in ["новая станция", "рассчитать станцию", "новый расчёт"]):
            mode = "new_station"
            st.session_state.station_profile = new_profile()
        elif any(w in q for w in CALC_TRIGGER_WORDS) and detect_calc_formula(user_input)[0]:
            mode = "calc"
        else:
            mode = "analyst"
        st.session_state.mode = mode

    # ── РЕЖИМ: НОВАЯ СТАНЦИЯ ──────────────────────────────
    if mode == "new_station":
        profile = st.session_state.station_profile or new_profile()
        step = get_next_question(profile)

        # Кнопка "Пропустить"
        if skipped:
            if step is None:
                st.session_state.history.append(("Пропустить", "Все вопросы уже отвечены.", "new_station"))
            elif step.get("default") is not None:
                profile[step["key"]] = step["default"]
                profile = apply_auto_fields(profile)
                next_step = get_next_question(profile)
                if next_step:
                    answer = f"Пропущено, использую значение по умолчанию: **{step['default']}**\n\n{format_question(next_step)}"
                else:
                    answer = (f"Пропущено. Все данные собраны.\n\n"
                              f"{profile_summary_text(profile)}\n\n"
                              "Напишите что угодно для запуска расчёта.")
                st.session_state.station_profile = profile
                st.session_state.history.append(("Пропустить", answer, "new_station"))
            else:
                answer = f"⚠️ Этот вопрос обязательный, пропустить нельзя.\n\n{format_question(step)}"
                st.session_state.history.append(("Пропустить", answer, "new_station"))
            st.rerun()

        # Кнопка "Отправить"
        if step is None:
            apply_defaults(profile)
            try:
                results = run_full_calculation(profile)
                st.session_state.station_results = results
                st.session_state.station_context = build_gigachat_context(profile, results)
                st.session_state.station_profile = profile
                st.session_state.mode = "station_qa"
                answer = build_report_markdown(profile, results)
                answer += "\n\n---\n✅ **Расчёт завершён.** Теперь вы можете задавать вопросы по станции."
            except Exception as e:
                answer = f"❌ Ошибка расчёта: {e}"
            st.session_state.history.append((user_input, answer, "new_station"))
        else:
            profile = apply_auto_fields(profile)
            val, err = parse_user_answer(step, user_input)
            if err:
                answer = f"⚠️ {err}\n\n{format_question(step)}"
            else:
                profile[step["key"]] = val
                profile = apply_auto_fields(profile)
                next_step = get_next_question(profile)
                if next_step:
                    answer = f"✓ Принято.\n\n{format_question(next_step)}"
                else:
                    answer = ("✓ Принято. Все данные собраны.\n\n"
                              f"{profile_summary_text(profile)}\n\n"
                              "Нажмите Отправить ещё раз для запуска расчёта.")
            st.session_state.station_profile = profile
            st.session_state.history.append((user_input, answer, "new_station"))

    # ── РЕЖИМ: Q&A ПО РАССЧИТАННОЙ СТАНЦИИ ───────────────
    elif mode == "station_qa":
        context = st.session_state.station_context or ""
        sys_prompt = SYSTEM_PROMPT_STATION + f"\n\n{context}"
        with st.spinner("Думаю..."):
            try:
                answer = call_gigachat(sys_prompt, user_input, st.session_state.history)
            except Exception as e:
                answer = f"❌ Ошибка: {e}"
        st.session_state.history.append((user_input, answer, "new_station"))

    # ── РЕЖИМ: КАЛЬКУЛЯТОР ────────────────────────────────
    elif mode == "calc":
        if st.session_state.calc_state is not None:
            cs = st.session_state.calc_state
            params = cs["data"]["params"]
            collected = cs["collected_params"]
            remaining = [p for p in params if p not in collected]
            cur_param = remaining[0]
            nums = re.findall(r"[-+]?\d*\.?\d+", user_input.replace(",", "."))
            if nums:
                collected[cur_param] = float(nums[0])
                remaining = [p for p in params if p not in collected]
                if not remaining:
                    try:
                        result = cs["data"]["calc"](**collected)
                        answer = (f"📐 **{cs['data']['formula']}**\n"
                                  f"Параметры: {collected}\n\n"
                                  f"✅ **Результат: {result:.3f} {cs['data']['units']}**")
                    except Exception as e:
                        answer = f"❌ Ошибка расчёта: {e}"
                    st.session_state.calc_state = None
                else:
                    next_p = remaining[0]
                    answer = f"✓ Принято. Введите: **{cs['data']['prompts'][next_p]}**"
            else:
                answer = "⚠️ Введите числовое значение."
            st.session_state.history.append((user_input, answer, "calc"))
        else:
            calc_key, calc_data = detect_calc_formula(user_input)
            if calc_key:
                nums = [float(n) for n in re.findall(r"[-+]?\d*\.?\d+", user_input.replace(",", "."))]
                params = calc_data["params"]
                if len(nums) >= len(params):
                    collected = dict(zip(params, nums[:len(params)]))
                    try:
                        result = calc_data["calc"](**collected)
                        answer = (f"📐 **{calc_data['formula']}**\n"
                                  f"Параметры: {collected}\n\n"
                                  f"✅ **Результат: {result:.3f} {calc_data['units']}**")
                    except Exception as e:
                        answer = f"❌ Ошибка: {e}"
                else:
                    first_p = params[0]
                    st.session_state.calc_state = {
                        "key": calc_key,
                        "data": calc_data,
                        "collected_params": {},
                    }
                    answer = (f"📐 Расчёт: **{calc_data['formula']}**\n\n"
                              f"Введите: **{calc_data['prompts'][first_p]}**")
            else:
                answer = ("Не распознана формула. Попробуйте написать, например:\n"
                          "«рассчитай ННЗТ для газовой станции» или «посчитай НЭЗТ уголь»\n\n"
                          "Доступные расчёты:\n" +
                          "\n".join(f"- {d['label']}" for d in CALC_FORMULAS.values()))
            st.session_state.history.append((user_input, answer, "calc"))

    # ── РЕЖИМ: АНАЛИТИК ───────────────────────────────────
    else:
        context_parts = []
        if excel_data:
            ec = find_relevant_excel(excel_data, user_input)
            if ec:
                context_parts.append(f"📊 ДАННЫЕ ПО СТАНЦИЯМ:\n{ec}")
        if collection is not None:
            kc = search_kb(user_input, collection, st_model)
            if kc:
                context_parts.append(f"📚 ИЗ БАЗЫ ЗНАНИЙ:\n{kc}")
        if st.session_state.station_context:
            context_parts.append(st.session_state.station_context)
        user_message = (user_input + "\n\n" + "\n\n".join(context_parts)
                        if context_parts else user_input)
        source = ("both" if len(context_parts) > 1
                  else "excel" if excel_data and find_relevant_excel(excel_data, user_input)
                  else "knowledge")
        with st.spinner("Думаю..."):
            try:
                answer = call_gigachat(SYSTEM_PROMPT_ANALYST, user_message,
                                       st.session_state.history)
            except Exception as e:
                answer = f"❌ Ошибка: {e}"
        st.session_state.history.append((user_input, answer, source))

    st.rerun()
